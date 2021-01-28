#!/usr/bin/env python
# -*- coding:utf-8 -*-

__author__ = 'shouke'

from datetime import datetime
from collections import OrderedDict
import shortuuid
import copy
import os
import json
import logging
import re
import xmind

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from django.conf import settings
from django.db import transaction
from django.core.files.uploadhandler import TemporaryFileUploadHandler
from django.utils import timezone
from django.http import FileResponse
from django.utils.http import urlquote


from backend.models import TestcaseSuite
from backend.models import SprintTestcase
from backend.models import BaseTestcase
from backend.models import SprintTestplanTestcase
from backend.models import SprintTestplan
from backend.models import TestcaseAttachment
from backend.models import SprintBaseCaseSuiteRelation
from backend.models import NormalFileDownload
from backend.models import TestPhase
from backend.serializers import SprintTestcaseSerializer
from backend.serializers import BaseTestcaseSerializer
from backend.serializers import SprintTestplanTestcaseSerializer
from backend.serializers import TestcaseSuiteSerializer
from backend.tester_view.testcase.testcase_suite_utils import TestSuiteStaticUtilsClass
from backend.utils.utils import mkdirs_in_batch
from backend.conf.config import MAX_CASES_NUM_XMIND_EXPORT, MAX_CASES_NUM_EXCEL_EXPORT, XMIND_TEMPLATE_FILE_PATH
from backend.tester_view.testplan.testplan_testcase_views import TestplanTestcaseResultAPIView


logger = logging.getLogger('mylogger')

class TestcaseListAPIView(APIView):
    '''
    测试视图-测试用例管理，查询测试用例表数据，批量删除测试用例
    '''

    def get(self, request, format=None):
        '''查询列表数据'''

        result = {}
        try:
            params =  request.GET
            start_time = params.get('startTime') # 创建起时时间
            end_time = params.get('endTime')  # 创建结束时间
            name = params.get('name') # 用例名称
            priority = params.get('priority')
            execution_phase = params.get('executionPhase')
            execution_method = params.get('executionMethod')
            executed_each_sprint = params.get('executedEachSprint')
            show_history_only = params.get('showHistoryOnly')
            recursive = int(params.get('recursive'))
            creater = params.get('creater')
            updater = params.get('updater')
            tag = params.get('tag')
            suite_type = params.get('suiteType')
            suite_id = int(params.get('suiteId'))
            sprint_id = params.get('sprintId')
            # product_id = params.get('product_id')
            page_size = int(params.get('pageSize'))
            page_no = int(params.get('pageNo'))
            sort = params.get('sort')
            testplan_id = params.get('testplanId')

            if (suite_type == 'sprint'):
                case_table_model = 'SprintTestcase'
                case_serializer = 'SprintTestcaseSerializer'
            elif suite_type == 'base':
                case_table_model = 'BaseTestcase'
                case_serializer = 'BaseTestcaseSerializer'
            else:
                result['msg'] =  '获取失败,测试所属套件类型(%s)不为base、sprint' % suite_type
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            if sort:
                sort_list = sort.split(',')
            else:
                sort_list = ['-id']

            startIndex = (page_no - 1) * page_size
            endIndex = startIndex + page_size

            filters = {'is_delete':0}
            if start_time:
                start_datetime = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__gte'] = start_datetime
            if end_time:
                end_datetime = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__lte'] = end_datetime

            if name:
                filters['name__startswith'] = name

            if priority:
                priority_list = priority.split(',')
                filters['priority__in'] = priority_list

            if execution_method:
                filters['execution_method'] = execution_method

            where_list = []
            if not executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s", execution_phase)' % execution_phase
                where_list.append(temp)
            elif executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s/%s", executed_each_sprint)' % (execution_phase, executed_each_sprint)
                where_list.append(temp)
            else:
                temp  = 'locate("/%s", executed_each_sprint)' % executed_each_sprint
                where_list.append(temp)

            if show_history_only == '1' and sprint_id:
                temp = 'sprint_id < %s OR sprint_id is null' % sprint_id
                where_list.append(temp)

            if creater:
                filters['creater_name'] = creater

            if updater:
                filters['updater_name'] = updater

            if tag:
                temp = 'find_in_set("%s", tags)' % tag
                where_list.append(temp)

            if recursive:
                suite_id_list = []
                suite_id_list.append(suite_id)

                # 获取子测试套件id
                query_sql = 'SELECT id FROM tb_testcase_suite AS t ' \
                           'WHERE FIND_IN_SET(%s, t.all_upper_node_ids) AND is_delete=0' % suite_id
                query_result = TestcaseSuite.objects.raw(query_sql)
                sub_suite_id_list = []
                for item in query_result:
                   sub_suite_id_list.append(int(item.id))
                suite_id_list.extend(sub_suite_id_list)
            else:
                suite_id_list = [suite_id]
            if suite_id_list:
                filters['suite_id__in'] = suite_id_list
            else:
                result['msg'] =  '获取成功'
                result['success'] =  True
                result['data'] = {}
                result['data']['rows'] = []
                result['data']['total'] = 0
                return Response(result, status.HTTP_200_OK)


            # 获取套件关联的测试用例
            if where_list:
                testcases = globals()[case_table_model].objects.filter(**filters).extra(where=where_list).order_by(*sort_list)[startIndex:endIndex]
                total = globals()[case_table_model].objects.filter(**filters).extra(where=where_list).count()
            else:
                testcases = globals()[case_table_model].objects.filter(**filters).order_by(*sort_list)[startIndex:endIndex]
                total = globals()[case_table_model].objects.filter(**filters).count()

            temp_dict = OrderedDict()
            temp_guid_list = [] # 存放获取的用例guid
            testplan_related_case_guids = [] # 存放计划关联的用例的guid
            if testcases:# 查找到测试用例
                testcases = globals()[case_serializer](testcases, many=True).data
                for testcase in testcases:
                    testcase['suite_type'] = suite_type
                    if testcase['tags']:
                        testcase['tags'] = testcase['tags'].split(',')
                    else:
                        testcase['tags'] = []
                    if testcase['execution_phase']:
                        testcase['execution_phase'] = testcase['execution_phase'].split(',')
                    else:
                        testcase['execution_phase'] = []
                    testcase['children'] = []

                    suite_id = testcase.get('suite_id')
                    if suite_id not in temp_dict:
                        suite_path = TestSuiteStaticUtilsClass.get_suite_path(suite_id)
                        suite_path = '/' + suite_path
                        guid = shortuuid.uuid()
                        temp_dict[suite_id] = {
                            'id':'',
                            'guid': guid,
                            'custom_no':'',
                            'suite_id':suite_id,
                            'sprint_id':'',
                            # 'product_id':product_id,
                            'parent_id': '',
                            'name':suite_path,
                            'priority': '',
                            'execution_phase':'',
                            'execution_method':'',
                            'executed_each_sprint':'',
                            'tags':'',
                            'desc':'',
                            'precondition':'',
                            'postcondition': '',
                            'creater_id': -1,
                            'creater_name':'',
                            'updater_id':-1,
                            'updater_name':'',
                            'children':[]
                        }

                    testcase['suitePath'] = temp_dict[suite_id]['name']
                    temp_dict[suite_id]['children'].append(testcase)
                    temp_guid_list.append(testcase.get('guid'))

                if temp_guid_list and testplan_id:
                    testplan_testcase_guids = SprintTestplanTestcase.objects.filter(testplan_id=testplan_id, guid__in=temp_guid_list).values_list('guid', flat=True)
                    if testplan_testcase_guids:
                        testplan_related_case_guids = testplan_testcase_guids

                testcases = []
                for key, value in temp_dict.items():
                    testcases.append(value)
            else:
                testcases = []
                total = 0

            result['msg'] =  '获取成功'
            result['success'] =  True
            result['data'] = {}
            result['data']['rows'] = testcases
            result['data']['total'] = total
            result['data']['testplanRelatedCaseGuids'] = testplan_related_case_guids
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


    # 批量删除列表数据
    def delete(self, request, format=None):
        try:
            data = request.data
            guids = data.get('guids')
            suite_type = data.get('suite_type')
            del_base_cases_cascade = data.get('del_base_cases_cascade') # 是否级联删除基线测试用例 字符串类型 1 级联删除 2 不删除


            result = {}
            try:
                with transaction.atomic():
                    if suite_type == 'base':
                        BaseTestcase.objects.filter(guid__in=guids).update(is_delete=1)
                        TestcaseAttachment.objects.filter(case_guid__in=guids).update(is_delete=1)
                    elif suite_type == 'sprint':
                        SprintTestcase.objects.filter(guid__in=guids).update(is_delete=1)
                        TestcaseAttachment.objects.filter(case_guid__in=guids).update(is_delete=1)

                        if del_base_cases_cascade == "1":
                            # 同步删除基线测试用例及关联的附件
                            BaseTestcase.objects.filter(sprint_testcase_guid__in=guids).update(is_delete=1)
                            TestcaseAttachment.objects.filter(sprint_case_guid__in=guids).update(is_delete=1)
                    else:
                        result['msg'] =  '删除失败,待删除测试用例所属套件类型只能为sprint、base'
                        result['success'] =  False
                        return Response(result, status.HTTP_400_BAD_REQUEST)

                    result['msg'] =  '删除成功'
                    result['success'] =  True
                    return Response(result, status.HTTP_200_OK)
            except Exception as e:
                result['msg'] =  '%s' % e
                result['success'] =  False
                return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

    # 批量修改用例
    def patch(self, request, format=None):
        result = {}
        try:
            data = request.data
            target_cases = data.get('target_cases')
            case_data = {}
            case_item_list = list(data.keys())
            case_item_list.remove('target_cases')
            case_item_list.remove('suite_id')
            for item in case_item_list:
                case_data[item] = data.get(item)
            case_data['updater_id'] = request.user.id
            case_data['updater_name'] = request.user.name
            for target_case in target_cases:
                suite_id = target_case['suite_id']
                suite_type = target_case['suite_type'].lower()
                guid = target_case['guid']
                testplan_id = target_case.get('testplan_id')
                detail_suite_type = target_case.get('detail_suite_type')
                if detail_suite_type:
                    case_data['detail_suite_type'] = detail_suite_type
                TestcaseAPIView.modify_testcase(suite_type, suite_id, guid, case_data, testplan_id)
            result['msg'] =  '修改成功'
            result['success'] =  True
            result['data']=[]
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


    @staticmethod
    def import_system_excel_testcases(template_file, product_id, creater_id, creater_name, test_phases_dict):
        try:
            table_field_list = ['casePath', 'name', 'priority', 'execution_phase', 'execution_method',
                                'precondition', 'action', 'expection', 'postcondition', 'tags', 'desc', 'guid']
            work_book = load_workbook(template_file)
            import_failure_info = []  # 存放导入失败的用例相关信息
            root_suite_info_dict = {} # 存放根套件相关信息
            case_path_father_suite_id_map = {} # 存放用例路径和用例直接上级套件ID的映射关系
            for sheet in work_book:
                line = 0 # 记录当前解析行
                for row in sheet.rows:
                    if line == 0: # 跳过第一行数据
                        line += 1
                        continue
                    line += 1
                    col_index = 0
                    row_data_dict = {}
                    for cell in row:
                        if cell.value is None:
                            row_data_dict[table_field_list[col_index]] = ''
                        else:
                            row_data_dict[table_field_list[col_index]] = str(cell.value).strip()
                        col_index += 1
                        if col_index == len(table_field_list):
                            break

                    fail_desc = ''
                    case_path =  row_data_dict['casePath'].strip('/').strip('\\')
                    if not case_path:
                        fail_desc = '用例路径为空'
                    elif case_path.count('/') < 1:
                        fail_desc = '用例路径只包含根测试集,用例路径至少要包含1个子测试集'
                    else:
                        if case_path not in case_path_father_suite_id_map:
                            suite_name_list = case_path.split('/')
                            root_suite_name = suite_name_list[0]
                            if root_suite_name not in root_suite_info_dict:
                                result = TestSuiteStaticUtilsClass.get_root_suite_info_by_name_for_product(root_suite_name, product_id)
                                if result['success']:
                                    root_suite_info_dict[root_suite_name] = result['root_suite_info']
                                else:
                                    fail_desc = result['reason']

                            if not fail_desc:
                                parent_suite_id =  root_suite_info_dict[root_suite_name].get('id') # 待创建套件的父级套件id
                                parent_suite_type = root_suite_info_dict[root_suite_name].get('type') # 待创建套件的父级套件类型
                                parent_suite_sprint_id = root_suite_info_dict[root_suite_name].get('sprint_id') # 待创建套件的父级套件关联迭代id
                                parent_suite_product_id =  root_suite_info_dict[root_suite_name].get('product_id') # 待创建套件的父级套件关联的产品id
                                result = TestSuiteStaticUtilsClass.create_sub_testsuite_by_path(case_path, parent_suite_type, parent_suite_id, parent_suite_sprint_id, parent_suite_product_id, creater_id)
                                if result['success']:
                                    parent_suite_id = result['tail_suite_id']
                                    case_path_father_suite_id_map[case_path] = parent_suite_id
                                else:
                                    fail_desc = result['reason']
                        else:
                            parent_suite_id = case_path_father_suite_id_map[case_path]

                    if not row_data_dict['name']:
                        fail_desc += '&用例名称为空'
                        continue # 用例名称为空，当做备注处理

                    if not row_data_dict['priority']:
                        fail_desc += '&优先级为空'
                    elif row_data_dict['priority'].upper() not in ['P1', 'P2', 'P3', 'P4']:
                        fail_desc += '&优先级填写错误'

                    final_execution_phase = ''
                    final_executed_each_sprint = ''
                    if not row_data_dict['execution_phase']:
                        fail_desc += '&执行阶段不能为空'
                    else:
                        row_data_dict['execution_phase'] = row_data_dict['execution_phase'].replace('，', ',').strip(',').replace(' ', '')
                        temp_execution_phase_list = row_data_dict['execution_phase'].split(',')
                        temp_execution_phase_set = set()
                        temp_executed_each_sprint_set = set()
                        executed_each_sprint_map = {'是':'Y', '否':'N'}
                        for item in temp_execution_phase_list:
                            temp_phase_list = item.strip('/').split('/')
                            execution_phase = ''.join(temp_phase_list[0:1])
                            executed_each_sprint = ''.join(temp_phase_list[1:2]) if temp_phase_list[1:2] else '否'
                            temp_executed_each_sprint_set.add(executed_each_sprint)
                            temp_execution_phase_set.add(execution_phase)
                            final_execution_phase += str(test_phases_dict.get(execution_phase)) + ','
                            final_executed_each_sprint += str(test_phases_dict.get(execution_phase)) + '/' \
                                                          + str(executed_each_sprint_map.get(executed_each_sprint)) + ','

                        final_execution_phase = final_execution_phase.strip(',')
                        final_executed_each_sprint = final_executed_each_sprint.strip(',')
                        temp_set = temp_execution_phase_set - set(list(test_phases_dict.keys()))
                        if temp_set:
                            fail_desc +='&执行阶段不存在：%s' % ','.join(list(temp_set))
                        temp_set = temp_executed_each_sprint_set - {'是', '否'}
                        if temp_set:
                            fail_desc +='&是否每次迭代都执行不存在：%s' % ','.join(list(temp_set))
                    if not row_data_dict['execution_method']:
                        fail_desc += '&执行方式不能为空'
                    elif row_data_dict['execution_method'] not in ['手工', '自动化']:
                        fail_desc += '&执行方式填写错误'
                    else:
                        execution_method_map = {'手工':'handwork', '自动化':'automation'}
                        row_data_dict['execution_method'] = execution_method_map[row_data_dict['execution_method']]

                    if fail_desc:
                        fail_desc = fail_desc.strip('&')
                        import_failure_info.append('所在表格：%s 所在行：%s 失败描述：%s<br>' %  (sheet.title, str(line), fail_desc))
                        continue

                    testcase_data = {}
                    testcase_data['guid'] = row_data_dict['guid']
                    testcase_data['suite_id'] = parent_suite_id
                    testcase_data['sprint_id'] = parent_suite_sprint_id if parent_suite_sprint_id != -1 else None
                    testcase_data['custom_no'] = None
                    testcase_data['name'] = row_data_dict['name']
                    testcase_data['priority'] = row_data_dict['priority'].upper()
                    testcase_data['execution_phase'] = final_execution_phase
                    testcase_data['execution_method'] = row_data_dict['execution_method']
                    testcase_data['executed_each_sprint'] = final_executed_each_sprint
                    testcase_data['precondition'] = row_data_dict['precondition']
                    testcase_data['steps'] = []

                    if row_data_dict['action']:
                        action_list = row_data_dict['action'].strip('\n').split('\n')
                        temp_str = action_list[0].strip()
                        result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', temp_str)
                        if result: # 遇到新步骤
                            temp_str =  temp_str.lstrip(result[0])
                        row_data_dict['action'] = []
                        for i in range(1, len(action_list)):
                            action = action_list[i].strip()
                            result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', action)
                            if result: # 遇到新步骤
                                row_data_dict['action'].append(temp_str)
                                temp_str =  action.lstrip(result[0])
                            else:
                                temp_str = temp_str + '\n' + action
                        row_data_dict['action'].append(temp_str)
                    else:
                        row_data_dict['action'] = []
                    if row_data_dict['expection']:
                        expection_list = row_data_dict['expection'].strip('\n').split('\n')
                        temp_str = expection_list[0].strip()
                        result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', temp_str)
                        if result: # 遇到新步骤
                            temp_str =  temp_str.lstrip(result[0])
                        row_data_dict['expection'] = []
                        for i in range(1, len(expection_list)):
                            expection = expection_list[i].strip()
                            result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', expection)
                            if result: # 遇到新步骤
                                row_data_dict['expection'].append(temp_str)
                                temp_str = expection.lstrip(result[0])
                            else:
                                temp_str = temp_str + '\n' + expection
                        row_data_dict['expection'].append(temp_str)
                    else:
                        row_data_dict['expection'] = []
                    for i in range(0, max(len(row_data_dict['action']), len(row_data_dict['expection']))):
                        action = ''.join(row_data_dict['action'][i:i+1])
                        expection = ''.join(row_data_dict['expection'][i:i + 1])
                        testcase_data['steps'].append({'action': action, 'expection':expection})

                    testcase_data['postcondition'] = row_data_dict['postcondition']
                    testcase_data['tags'] = row_data_dict['tags'].replace(',', ',').strip(',').replace(' ', '')
                    testcase_data['desc'] = row_data_dict['desc']
                    testcase_data['guid'] = row_data_dict['guid']
                    testcase_data['creater_id'] = creater_id
                    testcase_data['creater_name'] = creater_name
                    testcase_data['updater_id'] = creater_id
                    testcase_data['updater_name'] = creater_name
                    testcase_data['product_id'] = parent_suite_product_id
                    testcase_data['is_delete'] = 0
                    TestcaseListAPIView.save_case_data_to_db(testcase_data, testcase_data['suite_id'], parent_suite_type, sheet.title, line)
        except Exception as e:
            logger.error('导入系统模板用例出错：%s' % e)
            return {'msg':'导入系统模板用例出错:%s（部分用例可能已导入）' % e, 'success':False, 'status_code':status.HTTP_500_INTERNAL_SERVER_ERROR}

        if import_failure_info:
            import_failure_info = '<br/>'.join(import_failure_info)
            msg = '部分用例导入失败：<br/>%s' % import_failure_info
        else:
            msg = '用例全部导入成功'
        return {'msg':msg , 'success':True, 'status_code':status.HTTP_200_OK}


    @staticmethod
    def import_xmind_sheet_testcases(sheet_name, topic_data, topic_id_path_node_map, root_suite_path, root_suite_name, root_suites_info_dict, case_path_father_suite_id_map,
                                     execution_phase, creater_id, creater_name, import_failure_info, father_topic_id=None, topic_id_path_map={}):
        topic_id = topic_data.get('id')
        topic_markers = topic_data.get('markers')
        topic_label = topic_data.get('label')
        topic_title = topic_data.get('title')
        topic_comment = topic_data.get('note')
        fail_desc = ''

        testcase_steps = []
        sub_topic_list = topic_data.get('topics')
        if not father_topic_id and not sub_topic_list:
            logger.warn('未找到用例')
        elif not sub_topic_list or 'priority-' in ''.join(topic_markers): # 当前主题为最后一个主题、有优先级标记的主题当作测试用例对待
            if sub_topic_list: # 存在用例步骤
                for action_topic in sub_topic_list:
                    action =  action_topic.get('title')
                    if not action: # 操作步骤为空，跳过
                        continue
                    result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', action) # 去掉前面的步骤编号 1、2、3 、
                    if result: # 遇到新步骤
                        action =  action.lstrip(result[0])
                    expection_topic_list = action_topic.get('topics')
                    expection_str = ''
                    if expection_topic_list:
                        for expection_topic in expection_topic_list:
                            expection = expection_topic.get('title')
                            expection_str += (expection or '') + '\n'
                        expection_str = expection_str.rstrip('\n')
                    testcase_steps.append({'action': action, 'expection':expection_str})
            if father_topic_id and father_topic_id not in topic_id_path_map:
                father_topic_list = topic_id_path_node_map[father_topic_id]
                temp_path = ''
                for item in father_topic_list[1:]:
                    item_title = item.get('title').strip()
                    if item_title:
                        temp_path = temp_path + '/' + item_title
                temp_path = temp_path.lstrip('/')
                case_path = root_suite_path + '/' + temp_path
                case_path = case_path.rstrip('/')
                topic_id_path_map[father_topic_id] = case_path
            elif father_topic_id in topic_id_path_map:
                case_path = topic_id_path_map[father_topic_id]
            else:
                case_path = ''

            parent_suite_type = root_suites_info_dict[root_suite_name].get('type') # 待创建套件的父级套件类型
            parent_suite_sprint_id = root_suites_info_dict[root_suite_name].get('sprint_id') # 待创建套件的父级套件关联迭代id
            parent_suite_product_id = root_suites_info_dict[root_suite_name].get('product_id') # 待创建套件的父级套件关联的产品id
            if not case_path:
                fail_desc = '用例路径为空'
            elif case_path.count('/') < 1:
                fail_desc = '用例路径只包含根测试集,用例路径至少要包含1个子测试集'
            else:
                if case_path not in case_path_father_suite_id_map:
                    parent_suite_id =  root_suites_info_dict[root_suite_name].get('id') # 待创建套件的父级套件id
                    result = TestSuiteStaticUtilsClass.create_sub_testsuite_by_path(case_path, parent_suite_type, parent_suite_id, parent_suite_sprint_id, parent_suite_product_id, creater_id)
                    if result['success']:
                        parent_suite_id = result['tail_suite_id']
                        case_path_father_suite_id_map[case_path] = parent_suite_id
                    else:
                        fail_desc = result['reason']
                else:
                    parent_suite_id = case_path_father_suite_id_map[case_path]

            if not topic_title:
                fail_desc += '&用例名称为空'

            if fail_desc:
                fail_desc = fail_desc.strip('&')
                import_failure_info.append('所在画布：%s 失败用例：%s/%s 失败原因：%s<br/>' %  (sheet_name, case_path, topic_title, fail_desc))
                return

            if not topic_comment:
                topic_comment = ''

            priority = 'P3'
            for item in topic_markers:
                if 'priority-' in item:
                    priority = 'P%s' %  item.lstrip('priority-')
                    break
            testcase_data = {}
            testcase_data['guid'] = None
            testcase_data['suite_id'] = parent_suite_id
            testcase_data['sprint_id'] = parent_suite_sprint_id if parent_suite_sprint_id != -1 else None
            testcase_data['custom_no'] = None
            testcase_data['name'] = topic_title
            testcase_data['priority'] = priority
            testcase_data['execution_phase'] = execution_phase
            testcase_data['execution_method'] = 'handwork'
            executed_each_sprint = '%s/N' % execution_phase
            testcase_data['executed_each_sprint'] = executed_each_sprint
            testcase_data['precondition'] = ''
            testcase_data['steps'] = testcase_steps
            testcase_data['desc'] = topic_comment
            testcase_data['tags'] = ''
            testcase_data['product_id'] = parent_suite_product_id
            testcase_data['creater_id'] = creater_id
            testcase_data['creater_name'] = creater_name
            testcase_data['updater_id'] = creater_id
            testcase_data['updater_name'] = creater_name
            testcase_data['is_delete'] = 0
            TestcaseListAPIView.save_case_data_to_db(testcase_data, testcase_data['suite_id'], parent_suite_type, sheet_name)
        else: # 遇到的主题为测试集
            if father_topic_id:
                topic_id_path_node_map[topic_id] = copy.deepcopy(topic_id_path_node_map[father_topic_id])
                topic_id_path_node_map[topic_id].append({'markers': topic_markers, 'note': topic_comment, 'label':topic_label, 'title':topic_title})
            else:
                topic_id_path_node_map[topic_id] = [{'markers': topic_markers, 'note': topic_comment, 'label':topic_label, 'title':topic_title}]
            for sub_topic in sub_topic_list:
                TestcaseListAPIView.import_xmind_sheet_testcases(sheet_name, sub_topic, topic_id_path_node_map, root_suite_path, root_suite_name, root_suites_info_dict, case_path_father_suite_id_map,
                                                                 execution_phase, creater_id, creater_name, import_failure_info, topic_id, topic_id_path_map)


    @staticmethod
    def save_case_data_to_db(testcase_data, suite_id, suite_type, sheet_name, line=None):
        if line:
            logger.info('正在导入 %s 工作表 第 %s 行用例' % (sheet_name, str(line)))
        else:
            logger.info('正在导入用例，用例名称：%s 所在画布：%s' % (testcase_data['name'], sheet_name))
        testcase_data['steps'] = json.dumps(testcase_data['steps'])

        if suite_type == 'base':
            case_model = 'BaseTestcase'
        elif suite_type == 'sprint':
            case_model = 'SprintTestcase'
        else:
            return

        testcase = None
        if testcase_data['guid']: # 根据guid判断用例是否已存在
            testcase =  globals()[case_model].objects.filter(guid=testcase_data['guid'], suite_id=suite_id, is_delete=0).first()
        if not testcase: # 根据名称来判断用例是否已存在
            testcase =  globals()[case_model].objects.filter(name=testcase_data['name'], suite_id=suite_id, is_delete=0).first()
        if testcase:# 用例已存在，更新
            guid = testcase.guid
            del testcase_data['guid']
            del testcase_data['suite_id']
            TestcaseAPIView.modify_testcase(suite_type, suite_id, guid, testcase_data, None)
        else: # 否则，新增
            testcase_data['guid'] = shortuuid.uuid()
            TestcaseAPIView.new_testcase(suite_type, testcase_data, False, testcase_data['guid'], None)

    @staticmethod
    def import_system_xmind_testcases(template_file, product_id, creater_id, creater_name):
        try:
            result = {}
            root_suites_info_dict = {} # 存放根套件相关信息
            import_failure_info = []  # 存放导入失败的用例相关信息
            case_path_father_suite_id_map = {} # 存放用例路径和用例直接上级套件ID的映射关系

            test_phase_code = '' # 存放默认的测试用例执行阶段
            test_phases = TestPhase.objects.filter(is_delete=0).order_by('order', 'id').values('code', 'name', 'default')
            if not test_phases:
                result['msg'] =  '导入失败，未获取到测试阶段配置'
                result['success'] =  False
                return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                for test_phase in test_phases:
                    default = test_phase.get('default')
                    if default:
                        test_phase_code = test_phase.get('code')
                        break
                if not default:
                    test_phase_code = test_phases[0]['code']

            workbook = xmind.load(template_file)
            for sheet in workbook.getSheets():
                sheet_data_dict  = sheet.getData()
                sheet_name = sheet_data_dict.get('title')

                root_suite_name = '' # 根测试集名称
                root_suite_path = '' # 根测试集路径
                root_topic_data_dict = sheet_data_dict.get('topic')
                if root_topic_data_dict:
                    topic_title = root_topic_data_dict['title']
                    temp_list = topic_title.split('-')
                    root_suite_name = temp_list[0]
                    root_suite_path = ''
                    for item in temp_list:
                       item = item.strip()
                       if item:
                           root_suite_path += item + '/'
                    root_suite_path = root_suite_path.rstrip('/')

                if root_suite_name:
                    if root_suite_name not in root_suites_info_dict:
                        result = TestSuiteStaticUtilsClass.get_root_suite_info_by_name_for_product(root_suite_name, product_id)
                        if result['success']:
                             root_suites_info_dict[root_suite_name] = result['root_suite_info']
                        else:
                             import_failure_info.append('所在画布：%s 失败用例：所有用例，失败原因：%s<br/>' %  (sheet_name, result['reason']))
                             continue
                    else:
                        pass
                else:
                    import_failure_info.append('所在画布：%s 失败用例：所有用例，失败原因：%s<br/>' %  (sheet_name, '根测试集为空'))
                    continue

                topic_id_path_node_map = {} # # 存放主题id和主题所在路题结点的映射关系, {$id_value: [{topic_info_dict},{topic_info_dict},]}
                TestcaseListAPIView.import_xmind_sheet_testcases(sheet_name, root_topic_data_dict, topic_id_path_node_map, root_suite_path, root_suite_name, root_suites_info_dict,
                                                                case_path_father_suite_id_map, test_phase_code, creater_id, creater_name, import_failure_info)
        except Exception as e:
            logger.error('导入系统XMind模板用例出错：%s' % e)
            return {'msg':'导入系统XMind模板用例出错:%s（部分用例可能已导入）' % e, 'success':False, 'status_code':status.HTTP_500_INTERNAL_SERVER_ERROR}

        if import_failure_info:
            import_failure_info = '<br/>'.join(import_failure_info)
            msg = '部分用例导入失败：<br/>%s' % import_failure_info
        else:
            msg = '用例全部导入成功'
        return {'msg':msg , 'success':True, 'status_code':status.HTTP_200_OK}

    # 批量导入用例
    def post(self, request, format=None):
        result = {}
        try:
            request.upload_handlers = [TemporaryFileUploadHandler()]
            creater_id = request.user.id
            creater_name = request.user.name

            files = request.FILES
            file = files.get('file')

            if not file:
                result['msg'] =  '上传失败，未获取到文件'
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            test_phases = TestPhase.objects.filter(is_delete=0).values('code', 'name')
            test_phases_dict = {}
            if not test_phases:
                result['msg'] =  '导入失败，未获取到测试阶段配置'
                result['success'] =  False
                return  Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                for test_phase in test_phases:
                    test_phases_dict[test_phase.get('name')] = test_phase.get('code')

            data = request.POST
            product_id = data.get('product_id')
            case_template_type = data.get('case_template_type')
            temp_file_path = file.temporary_file_path()

            if case_template_type == 'SYSTEM_EXCEL_TESTCASE_TEPLATE':
                temp_result = TestcaseListAPIView.import_system_excel_testcases(temp_file_path, product_id, creater_id, creater_name, test_phases_dict)
            elif case_template_type == 'SYSTEM_XMIND_TESTCASE_TEPLATE':
                temp_result = TestcaseListAPIView.import_system_xmind_testcases(temp_file_path, product_id, creater_id, creater_name)
            else:
                result['msg'] = '导入失败，不支持的模板类型'
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            result['msg'] = temp_result['msg']
            result['success'] =  temp_result['success']
            status_code = temp_result['status_code'] # status.HTTP_200_OK
            return Response(result, status_code)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


class TestcaseAPIUtils:
    '''工具类，解决testcase_suite_views.PasteTestcasesCopiedAPIView和testcase_views.TestcaseAPIView相互引用问题'''

    @staticmethod
    def copy_case_attachments(to_case_guid, src_case_guid, type):
        '''
        复制测试用例附件到附件表
        '''
        filters = {'is_delete':0, 'case_guid':src_case_guid}
        attachments = TestcaseAttachment.objects.filter(**filters)
        for attachment in attachments:
            guid = shortuuid.uuid()
            case_guid = to_case_guid
            name = attachment.name

            file_path = attachment.file_path
            path, suffix = os.path.splitext(file_path)
            file_name = os.path.basename(file_path)
            time_str = path[-8:]
            target_file_name = shortuuid.uuid() + time_str + suffix
            file_relative_path = file_path.replace(file_name, target_file_name)
            file_absolute_path = settings.MEDIA_ROOT.rstrip('/') + file_relative_path # file_relative_path 以 / 打头

            # 生成文件
            src_file_absolute_path = settings.MEDIA_ROOT.rstrip('/') + file_path
            if not os.path.exists(src_file_absolute_path): # 源文件不存在，则结束,不做任何处理
                return
            target_file = open(file_absolute_path, 'wb')
            try:
                with open(src_file_absolute_path, 'rb') as f:
                    for content in f:
                        target_file.write(content)
            finally:
                target_file.close()

            file_path = file_relative_path
            creater_id = attachment.creater_id
            create_time = attachment.create_time
            is_delete = attachment.is_delete

            if type == 'sprint_to_base':
                obj = TestcaseAttachment(guid=guid, case_guid=case_guid, sprint_case_guid=src_case_guid, attachment_guid=attachment.guid, name=name, file_path=file_path, creater_id=creater_id, create_time=create_time, is_delete=is_delete)

            else:
                obj = TestcaseAttachment(guid=guid, case_guid=case_guid, sprint_case_guid=src_case_guid, name=name, file_path=file_path, creater_id=creater_id, create_time=create_time, is_delete=is_delete)
            obj.save()

            if type == 'base_to_sprint': # 更改基线用例附件和迭代用例附件的映射关系
                attachment.attachment_guid = obj.guid
                attachment.save()

    # 同步sprint测试用例数据到基线测试套件
    # data:需要同步的数据
    @staticmethod
    def sync_sprint_case_to_base(sprint_case_guid, sprint_case_suite_id, data, operate='new'):
        # 查找对应基线测试用例
        testcase = BaseTestcase.objects.filter(sprint_testcase_guid=sprint_case_guid, is_delete=0).first()
        if testcase: # 存在用例
            # 更新对应的基线用例
            serializer = BaseTestcaseSerializer(testcase, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        else: # 插入基线测试用例
            # 查找迭代测试套件关联的基线测试套件
            relation = SprintBaseCaseSuiteRelation.objects.filter(sprint_case_suite_id=sprint_case_suite_id).first()

            base_suite = None
            base_suite_id = None
            if relation: # 如果存在关系，就一定存在对应套件，因为删除套件时会自动删除关系
                base_suite = TestcaseSuite.objects.filter(id=relation.base_case_suite_id).first()
            else: # 判断用例关联的测试套件是否为根套件，如果是则用例要挂载至的目标基线套件默认为根套件
                sprint_suite = TestcaseSuite.objects.filter(id=sprint_case_suite_id).first()
                if sprint_suite and sprint_suite.parent_id == -1: # 用例关联的迭代测试套件为根测试套件，则默认同步创建的基线用例挂载的基线测试套件为根测试套件
                    base_suite =  TestcaseSuite.objects.filter(type='base', is_delete=0, product_id=sprint_suite.product_id, parent_id=-1, sprint_id=-1).first()
                else: # 递归创建基线对应的迭代测试套件
                    base_suite_id = TestSuiteStaticUtilsClass.sync_sprint_suite_to_base_recursive(TestcaseSuiteSerializer(sprint_suite).data)

            if base_suite:
                base_suite_id = base_suite.id

            if base_suite_id:
                data['guid'] = shortuuid.uuid()
                data['sprint_testcase_guid'] = sprint_case_guid
                if 'creater_id' not in data.keys(): # 修改迭代测试用例，发现基线测试用例不存在，此时data中并没包含创建人信息,设置更新人为创建人
                    data['creater_id'] = data.get('updater_id')
                    data['creater_name'] = data.get('updater_name')
                data['suite_id'] = base_suite_id

                # 查询迭代测试用例，获取用例信息（data可能是修改传递过来的数据，不一定是创建用例所需的完整数据,因为可能是局部修改，修改用例也需要同步基线，此时，如果基线用例不存在，也要创建用例）
                if operate == 'update':
                    sprint_case = SprintTestcase.objects.filter(guid=sprint_case_guid, is_delete=0).first()
                    if sprint_case:
                        data['custom_no'] = None
                        data['product_id'] = sprint_case.product_id
                        data['sprint_id'] = sprint_case.sprint_id
                        data['name'] = sprint_case.name
                        data['priority'] = sprint_case.priority
                        data['executed_each_sprint'] = sprint_case.executed_each_sprint
                        data['execution_method'] = sprint_case.execution_method
                        data['execution_phase'] = sprint_case.execution_phase
                        data['tags'] = sprint_case.tags
                        data['desc'] = sprint_case.desc
                        data['precondition'] = sprint_case.precondition
                        data['steps'] = sprint_case.steps
                        data['postcondition'] = sprint_case.postcondition
                        data['desc'] = sprint_case.desc
                        data['is_delete'] = 0
                    else: # 迭代用例不存在，则不作任何处理
                        pass
                base_testcase_serializer = BaseTestcaseSerializer(data=data)
                base_testcase_serializer.is_valid(raise_exception=True)
                base_testcase_serializer.save()

                base_case_guid = base_testcase_serializer.data.get('guid')
                # 拷贝迭代测试用例附件到测试附件表
                TestcaseAPIUtils.copy_case_attachments(base_case_guid, sprint_case_guid, 'sprint_to_base')

    @staticmethod
    def sync_plan_case_to_sprint(guid, suite_id, data):
        '''同步计划用例到迭代用例'''

        testcase = SprintTestcase.objects.filter(guid=guid).first()
        if testcase:
            try:
                with transaction.atomic():
                    serializer = SprintTestcaseSerializer(testcase, data=data, partial=True)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()

                    # 同步基线测试用例
                    TestcaseAPIUtils.sync_sprint_case_to_base(guid, suite_id, data, 'update')
            except Exception as e:
                raise Exception("%s" % e)
        else: # 用例不存在，啥都不操作
            pass

    @staticmethod
    def sync_plan_case_to_base(guid, data):
        '''更新计划用例到基线用例'''

        testcase = BaseTestcase.objects.filter(guid=guid).first()
        if testcase:
            try:
                serializer = BaseTestcaseSerializer(testcase, data=data, partial=True)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            except Exception as e:
                raise Exception("%s" % e)
        else: # 用例不存在，啥都不操作
            pass

class TestcaseAPIView(APIView):
    '''
    测试视图-测试用例管理，新增，修改，删除用例
    '''

    @staticmethod
    def new_testcase(suite_type, testcase_data, newcase_by_copy, target_case_guid, src_case_guid, plan_case_statistics=[]):
        '''
        新增测试用例
        '''

        creater_id = testcase_data.get('creater_id')
        creater_name = testcase_data.get('creater_name')

        if suite_type == 'base':
            case_serializer = 'BaseTestcaseSerializer'
        elif suite_type == 'sprint':
            case_serializer = 'SprintTestcaseSerializer'
        else:
            raise Exception('新增失败,请先点选待新建用例归属的测试集')
        serializer = globals()[case_serializer](data=testcase_data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            serializer.save()
            if newcase_by_copy:# 通过复制新建的用例
                type = ''
                if suite_type == 'base':
                    type = 'base_to_base'
                elif suite_type == 'sprint':
                    type = 'sprint_to_sprint'

                # 拷贝源用例的测试附件到新建测试用例下
                TestcaseAPIUtils.copy_case_attachments(target_case_guid, src_case_guid, type)

            if suite_type == 'sprint': # 同步基线测试用例
                TestcaseAPIUtils.sync_sprint_case_to_base(target_case_guid, testcase_data.get('suite_id'), copy.deepcopy(testcase_data))

            if testcase_data.get('testplan_id'): # 需要新增测试计划用例
                testcase_data['create_time'] = serializer.data.get('create_time')
                testcase_data['update_time'] = serializer.data.get('update_time')
                testcase_data['assigned_to_id'] = serializer.data.get('creater_id')
                testcase_data['assigned_to'] = testcase_data.get('creater_name')
                testcase_data['tester_id'] = None
                serializer = SprintTestplanTestcaseSerializer(data=testcase_data)
                serializer.is_valid(raise_exception=True)
                serializer.save()

                # 更新测试计划用例计数器
                testplan = SprintTestplan.objects.select_for_update().filter(id=testcase_data.get('testplan_id')).first()
                if not testplan:
                     raise Exception('未找到目标计划（目标计划ID：%s)' % testcase_data.get('testplan_id'))

                testplan.case_num_related = testplan.case_num_related + 1
                testplan.finish_time = None

                # 更新测试计划统计数据
                plan_case_statistics.append(TestplanTestcaseResultAPIView.update_testplan_info(testplan, testplan.status,  timezone.now(), creater_id, creater_name))
        return serializer.data


    @staticmethod
    def modify_testcase(suite_type, suite_id, testcase_guid, testcase_data, testplan_id):
        '''
        修改测试用例
        '''

        if suite_type == 'base':
            case_table_model = 'BaseTestcase'
            case_serializer = 'BaseTestcaseSerializer'
            filters = {'guid':testcase_guid}
        elif suite_type == 'sprint':
            case_table_model = 'SprintTestcase'
            case_serializer = 'SprintTestcaseSerializer'
            filters = {'guid':testcase_guid}
        elif suite_type == 'testplan': # 修改迭代测试计划关联的测试用例
            case_table_model = 'SprintTestplanTestcase'
            case_serializer = 'SprintTestplanTestcaseSerializer'
            filters = {'guid':testcase_guid, 'testplan_id':testplan_id}
        else:
            raise Exception('修改失败,测试用例所属套件类型只能为sprint、base、testplan')

        testcase = globals()[case_table_model].objects.filter(**filters).first()
        if testcase:
            try:
                serializer = globals()[case_serializer](testcase, data=testcase_data, partial=True)
                serializer.is_valid(raise_exception=True)

                with transaction.atomic():
                    serializer.save()
                    if suite_type == 'sprint': # 同步基线测试用例
                        TestcaseAPIUtils.sync_sprint_case_to_base(testcase_guid, suite_id, testcase_data, 'update')
                    elif suite_type == 'testplan':
                        detail_suite_type = testcase_data.get('detail_suite_type')
                        if detail_suite_type == 'sprint':
                            # 同步修改迭代测试用例基本信息
                            TestcaseAPIUtils.sync_plan_case_to_sprint(testcase_guid, suite_id, testcase_data)
                        elif detail_suite_type == 'base':
                            # 同步修改基线测试用例信息
                            TestcaseAPIUtils.sync_plan_case_to_base(testcase_guid, testcase_data)
                        else:
                            pass
                    return serializer.data
            except Exception as e:
                raise Exception('操作失败：%s' % e)
        else:
            raise Exception('操作失败,用例不存在')


     # 新增用例
    def post(self, request, format=None):
        result = {}
        try:
            data = request.data
            src_case_guid = data.get('guid')
            newcase_by_copy = False
            if src_case_guid:
                del data['guid']
                newcase_by_copy = True

            guid = shortuuid.uuid()
            data['guid'] = guid # 待新增用例的guid
            data['creater_id'] = request.user.id # 创建人id
            data['updater_id'] = data['creater_id'] # 更新人id
            data['creater_name'] = request.user.name
            data['updater_name'] = data['creater_name']
            data['is_delete'] = 0              # 用例是否删除
            suite_type = data.get('suite_type').lower()    # 用例测试套件类型
            # suite_id = data.get('suite_id')        # 待新增用例关联的测试套件id
            try:
                plan_case_statistics = [] # 存放测试计划用例统计数据
                serializer_data_for_case_new = TestcaseAPIView.new_testcase(suite_type, data, newcase_by_copy, guid, src_case_guid, plan_case_statistics)
                case_data = copy.deepcopy(serializer_data_for_case_new)
                if case_data['tags']:
                    case_data['tags'] = case_data['tags'].split(',')
                else:
                    case_data['tags'] = []
                if case_data['execution_phase']:
                    case_data['execution_phase'] = case_data['execution_phase'].split(',')
                else:
                    case_data['execution_phase'] = []
                case_data['suitePath'] = data.get('suite_path')
                if data.get('testplan_id'): # 新增测试计划用例
                    case_data['suiteType'] = 'testplan'
                    case_data['detailSuiteType'] = 'sprint'
                else:
                    case_data['suiteType'] = suite_type
                case_data['children'] = []
            except Exception as e:
                result['msg'] =  '新增失败：%s' % e
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            result['msg'] =  '新增成功'
            result['success'] =  True
            if len(plan_case_statistics):
                plan_case_statistics = plan_case_statistics[0]
            result['data'] =  {'case':case_data, 'plan_case_statistics':plan_case_statistics}
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

    # 修改用例
    def patch(self, request, format=None):
        result = {}
        try:
            data = request.data
            data['updater_id'] = request.user.id
            data['updater_name'] = request.user.name
            suite_type = data.get('suite_type').lower()
            guid = data.get('guid')
            suite_id = data.get('suite_id')
            testplan_id = data.get('testplan_id')


            if testplan_id:
                del data['testplan_id']
            if 'suite_path' in data:
                del data['suite_path']
            del data['guid']
            del data['case_id']
            del data['suite_id']
            del data['suite_type']

            case_serializer_data = TestcaseAPIView.modify_testcase(suite_type, suite_id, guid, data, testplan_id)

            result['msg'] =  '修改成功'
            result['success'] =  True
            data_updated = {}
            data_keys = list(data.keys())
            data_keys.append('update_time')
            for key in data_keys:
                data_updated[key] = case_serializer_data.get(key)
            result['data'] =  data_updated
            if 'tags' in data_updated:
                if data_updated['tags']:
                    data_updated['tags'] = data_updated['tags'].split(',')
                else:
                    data_updated['tags'] = []

            if 'execution_phase' in data_updated:
                if data_updated['execution_phase']:
                    data_updated['execution_phase'] = data_updated['execution_phase'].split(',')
                else:
                    data_updated['execution_phase'] = []
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

    # 逐条删除用例
    def delete(self, request, format=None):
        result = {}
        try:
            data = request.data
            guid = data.get('guid')
            suite_type = data.get('suite_type')
            del_base_cases_cascade = data.get('del_base_cases_cascade') # 是否级联删除基线测试用例 字符串类型 1 级联删除 2 不删除

            if suite_type == 'base':
                BaseTestcase.objects.filter(guid=guid).update(is_delete=1)
                TestcaseAttachment.objects.filter(case_guid=guid).update(is_delete=1)
            elif suite_type == 'sprint':
                try:
                    with transaction.atomic():
                        SprintTestcase.objects.filter(guid=guid).update(is_delete=1)
                        TestcaseAttachment.objects.filter(case_guid=guid).update(is_delete=1)

                        if del_base_cases_cascade == "1": # 否则 不做任何处理
                            # 同步删除基线测试用例
                            BaseTestcase.objects.filter(sprint_testcase_guid=guid).update(is_delete=1)
                            TestcaseAttachment.objects.filter(sprint_case_guid=guid).update(is_delete=1)
                except Exception as e:
                    result['msg'] =  '删除失败,%s' % e
                    result['success'] =  False
                    return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                result['msg'] =  '删除失败,待删除测试用例所属套件类型只能为sprint、base'
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)
            result['msg'] =  '删除成功'
            result['success'] =  True
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


class TestcaseExportAPIView(APIView):
    '''
    测试视图-测试用例管理，批量导出测试用例
    '''

    def get(self, request, format=None):
        '''查询列表数据'''

        result = {}
        try:
            params =  request.GET
            start_time = params.get('startTime') # 创建起时时间
            end_time = params.get('endTime')  # 创建结束时间
            name = params.get('name') # 用例名称
            priority = params.get('priority')
            execution_phase = params.get('executionPhase')
            execution_method = params.get('executionMethod')
            executed_each_sprint = params.get('executedEachSprint')
            show_history_only = params.get('showHistoryOnly')
            recursive = int(params.get('recursive'))
            creater = params.get('creater')
            updater = params.get('updater')
            tag = params.get('tag')
            suite_type = params.get('suiteType')
            suite_id = params.get('suiteId')
            sprint_id = params.get('sprintId')
            sort = params.get('sort')
            # testplan_id = params.get('testplanId')

            if (suite_type == 'sprint'):
                case_table_model = 'SprintTestcase'
            elif suite_type == 'base':
                case_table_model = 'BaseTestcase'
            else:
                result['msg'] =  '导出失败,测试所属套件类型(%s)不为base、sprint' % suite_type
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            test_phases = TestPhase.objects.filter(is_delete=0).values('code', 'name')
            test_phases_dict = {}
            if not test_phases:
                result['msg'] =  '导出失败，未获取到测试阶段配置'
                result['success'] =  False
                return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                for test_phase in test_phases:
                    test_phases_dict[test_phase.get('code')] = test_phase.get('name')

            if sort:
                sort_list = sort.split(',')
            else:
                sort_list = ['-id']

            filters = {'is_delete':0}
            if start_time:
                start_datetime = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__gte'] = start_datetime
            if end_time:
                end_datetime = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__lte'] = end_datetime

            if name:
                filters['name__startswith'] = name

            if priority:
                priority_list = priority.split(',')
                filters['priority__in'] = priority_list

            if execution_method:
                filters['execution_method'] = execution_method

            where_list = []
            if not executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s", execution_phase)' % execution_phase
                where_list.append(temp)
            elif executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s/%s", executed_each_sprint)' % (execution_phase, executed_each_sprint)
                where_list.append(temp)
            else:
                temp  = 'locate("/%s", executed_each_sprint)' % executed_each_sprint
                where_list.append(temp)

            if show_history_only == '1' and sprint_id:
                temp = 'sprint_id < %s OR sprint_id is null' % sprint_id
                where_list.append(temp)

            if creater:
                filters['creater_name'] = creater

            if updater:
                filters['updater_name'] = updater

            if tag:
                temp = 'find_in_set("%s", tags)' % tag
                where_list.append(temp)

            if recursive:
                suite_id_list = []
                suite_id_list.append(int(suite_id))

                # 获取子测试套件id
                query_sql = 'SELECT id FROM tb_testcase_suite AS t ' \
                           'WHERE FIND_IN_SET(%s, t.all_upper_node_ids) AND is_delete=0' % int(suite_id)
                # 暂不使用全文索引
                # query_sql = 'SELECT id FROM tb_testcase_suite AS t ' \
                #            'WHERE MATCH(all_upper_node_ids) AGAINST (%s IN BOOLEAN MODE) AND is_delete=0' % suite_id
                query_result = TestcaseSuite.objects.raw(query_sql)
                sub_suite_id_list = []
                for item in query_result:
                   sub_suite_id_list.append(int(item.id))
                suite_id_list.extend(sub_suite_id_list)
            else:
                suite_id_list = [int(suite_id)]
            if suite_id_list:
                filters['suite_id__in'] = suite_id_list
            else:
                result['msg'] =  '导出用例失败，获取测试用所属测试套件为空'
                result['success'] =  False
                return Response(result, status.HTTP_200_OK)

            # 获取套件关联的测试用例
            if where_list:
                testcases = globals()[case_table_model].objects.filter(**filters).extra(where=where_list).order_by(*sort_list)[0:MAX_CASES_NUM_EXCEL_EXPORT]
            else:
                testcases = globals()[case_table_model].objects.filter(**filters).order_by(*sort_list)[0:MAX_CASES_NUM_EXCEL_EXPORT]

            temp_dict = OrderedDict()
            work_book = Workbook()
            case_sheet = work_book.create_sheet(title='测试用例',index=0)
            case_sheet.column_dimensions['A'].width = 40
            case_sheet.column_dimensions['B'].width = 40
            case_sheet.column_dimensions['C'].width = 7
            case_sheet.column_dimensions['D'].width = 30
            case_sheet.column_dimensions['G'].width = 25
            for col in [ 'H', 'I', 'J', 'K', 'L']:
                case_sheet.column_dimensions[col].width = 50
            # 添加表头
            row = ['用例路径', '用例名称', '优先级', '执行阶段/每个迭代都执行？','执行方式','前置条件','测试步骤','预期结果','后置条件','标签','用例备注','系统用例ID']
            case_sheet.append(row)
            ord_a = ord('A')
            for item in range(ord_a, ord_a + len(row)+1):
                col = chr(item)
                cell =  col + '1'
                case_sheet[cell].font = Font(bold=True)
                case_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")

            for testcase in testcases:
                suite_id = testcase.suite_id
                if suite_id not in temp_dict:
                    suite_path = TestSuiteStaticUtilsClass.get_suite_path(suite_id)
                    suite_path = '/' + suite_path
                    temp_dict[suite_id] = {
                        'suitePath':suite_path,
                        'children':[]
                    }
                testcase.suitePath = temp_dict[suite_id]['suitePath']
                steps = json.loads(testcase.steps, object_pairs_hook=OrderedDict)

                action = ''
                expection = ''
                cnt = 1
                for step in steps:
                    result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', step['action'])
                    if result:
                        step['action'] = step['action'].lstrip(result[0]).strip()
                    result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', step['expection'])
                    if result:
                        step['expection'] = step['expection'].lstrip(result[0]).strip()
                    action += '%s、' % cnt + step['action'] + '\n'
                    expection +=  '%s、' % cnt + step['expection'] + '\n'
                    cnt += 1
                action = action.rstrip('\n')
                expection = expection.rstrip('\n')
                expection = expection.replace('\n\n', '\n')

                temp_execution_phase_list = testcase.executed_each_sprint.split(',')
                final_execution_phase = ''
                executed_each_sprint_map = {'Y':'是', 'N':'否'}
                for item in temp_execution_phase_list:
                    item_list = item.strip().replace(' ', '').split('/')
                    execution_phase = ''.join(item_list[:1])
                    executed_each_sprint = ''.join(item_list[1:2])
                    if execution_phase in test_phases_dict:
                        final_execution_phase += test_phases_dict[execution_phase] +  '/' + executed_each_sprint_map[executed_each_sprint] + ','
                final_execution_phase = final_execution_phase.rstrip(',')  # 形如 冒烟测试/是,预发布测试/否,系统测试/否,上线测试/是

                execution_method_map = {'handwork':'手工', 'automation':'自动化'}
                execution_method = execution_method_map[testcase.execution_method]
                row = [testcase.suitePath, testcase.name, testcase.priority, final_execution_phase, execution_method, testcase.precondition, action, expection, testcase.postcondition, testcase.tags, testcase.desc, testcase.guid]
                temp_dict[suite_id]['children'].append(row)

            for suite_id in temp_dict.keys():
                children = temp_dict[suite_id]['children']
                for row in children:
                    case_sheet.append(row)
            time_str = timezone.now().strftime('%Y%m%d%H%M%S%f')
            file_name =  shortuuid.uuid() + time_str + '.xlsx'
            file_dir = settings.MEDIA_ROOT.rstrip('/') + '/testcase/download'
            if not os.path.exists(file_dir):# 路径不存在
                if not mkdirs_in_batch(file_dir):
                    result['msg'] =  '导出用例失败，批量创建路径(%s)对应的目录失败'  % file_dir
                    result['success'] =  False
                    return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)

            file_absolute_path =  '%s/%s' % (file_dir, file_name)
            work_book.save(file_absolute_path)
            file = open(file_absolute_path, 'rb')
            file_response = FileResponse(file)
            file_response['Content-Type']='application/octet-stream'
            temp_file_name = '导出测试用例%s' %  time_str
            file_response["Access-Control-Expose-Headers"] = 'Content-Disposition'
            file_response['Content-Disposition']='attachment;filename={}.xlsx'.format(urlquote(temp_file_name))

            relative_file_path = '/testcase/download/%s' % file_name
            try:
                # 存储导出测试用例文件到数据库
                obj = NormalFileDownload(type='测试用例', name=temp_file_name, file_path=relative_file_path, creater=request.user.name, creater_id=request.user.id,  updater_id=request.user.id,  updater=request.user.name, is_delete=0)
                obj.save()
            except Exception as e:
                file.close()
                os.remove(file_absolute_path)
                raise Exception(e)
            return file_response
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)


class XMindTestcaseExportAPIView(APIView):
    '''
    测试视图-测试用例管理，导出测试用例为XMind
    '''
    #
    # @staticmethod
    # def create_xmind_case_suite_topic(suite_path, suite_node_dict, xmind_sheet):
    #     '''
    #     根据测试用例套件路径创建xmind结点(全部数据放一个sheet表中)
    #     suite_node_dict: 存放结点和用例套件对应关系
    #     '''
    #     head = os.path.normpath(suite_path)  # 去掉路径最右侧的 \\ 、/
    #     head = head.replace('\\', '/')  # 将所有的\\转为/，避免出现转义字符串
    #     temp_suite_path_list = [] # 临时存放没有对应xmind主题的测试套件所在全路径
    #     # tail = head
    #     # while head and tail != '':
    #     while '/' in head.lstrip('/'):
    #         temp_head = head
    #         head, tail = os.path.split(head)
    #         # if temp_head not in suite_node_dict and temp_head != '/':
    #         if temp_head not in suite_node_dict :
    #             temp_suite_path_list.append(temp_head)
    #     # if head == '/':
    #     if head.startswith('/') and head not in suite_node_dict:
    #         suite_node_dict[head] = xmind_sheet.getRootTopic()
    #         topic_name = head.lstrip('/')
    #         suite_node_dict[head].setTitle(topic_name)
    #         root_topic1 = xmind_sheet.getRootTopic()  # 获取画布中心主题，默认创建画布时会新建一个空白中心主题
    #         root_topic1.setTitle(topic_name)   # 设置初始名称
    #
    #     list_len = len(temp_suite_path_list)
    #     for i in range(0, list_len):
    #         path = temp_suite_path_list[list_len - 1 - i]
    #         head, tail = os.path.split(path)
    #         suite_node_dict[path] = suite_node_dict[head].addSubTopic()
    #         suite_node_dict[path].setTitle(tail)

    @staticmethod
    def create_xmind_case_suite_topic(suite_path, suite_node_dict, xmind_workbook):
        '''
        根据测试用例套件路径创建xmind结点
        suite_node_dict: 存放结点和用例套件对应关系
        '''
        head = os.path.normpath(suite_path)  # 去掉路径最右侧的 \\ 、/
        head = head.replace('\\', '/')  # 将所有的\\转为/，避免出现转义字符串
        temp_suite_path_list = [] # 临时存放没有对应xmind主题的测试套件所在全路径

        while head.count('/') > 3:
            if head not in suite_node_dict :
                temp_suite_path_list.append(head)
            head, tail = os.path.split(head)

        if head.count('/') in [2, 3] and head not in suite_node_dict:
            if suite_node_dict:
                xmind_sheet = xmind_workbook.createSheet()
            else:# 第一个sheet未被占用
                xmind_sheet = xmind_workbook.getPrimarySheet()
            suite_node_dict[head] = xmind_sheet.getRootTopic()
            suite_node_dict[head].setTitle(head.lstrip('/').replace('/', '-'))
            xmind_sheet.setTitle('-'.join(head.split('/')[2:]))

        list_len = len(temp_suite_path_list)
        for i in range(0, list_len):
            path = temp_suite_path_list[list_len - 1 - i]
            head, tail = os.path.split(path)
            suite_node_dict[path] = suite_node_dict[head].addSubTopic()
            suite_node_dict[path].setTitle(tail)

    def get(self, request, format=None):
        '''查询列表数据'''

        result = {}
        try:
            params =  request.GET
            start_time = params.get('startTime') # 创建起时时间
            end_time = params.get('endTime')  # 创建结束时间
            name = params.get('name') # 用例名称
            priority = params.get('priority')
            execution_phase = params.get('executionPhase')
            execution_method = params.get('executionMethod')
            executed_each_sprint = params.get('executedEachSprint')
            show_history_only = params.get('showHistoryOnly')
            recursive = int(params.get('recursive'))
            creater = params.get('creater')
            updater = params.get('updater')
            tag = params.get('tag')
            suite_type = params.get('suiteType')
            suite_id = params.get('suiteId')
            sprint_id = params.get('sprintId')
            sort = params.get('sort')
            # testplan_id = params.get('testplanId')

            if (suite_type == 'sprint'):
                case_table_model = 'SprintTestcase'
            elif suite_type == 'base':
                case_table_model = 'BaseTestcase'
            else:
                result['msg'] =  '获取失败,测试所属套件类型(%s)不为base、sprint' % suite_type
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            if sort:
                sort_list = sort.split(',')
            else:
                sort_list = ['-id']

            filters = {'is_delete':0}
            if start_time:
                start_datetime = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__gte'] = start_datetime
            if end_time:
                end_datetime = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                filters['create_time__lte'] = end_datetime

            if name:
                filters['name__startswith'] = name

            if priority:
                priority_list = priority.split(',')
                filters['priority__in'] = priority_list

            if execution_method:
                filters['execution_method'] = execution_method


            where_list = []
            if not executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s", execution_phase)' % execution_phase
                where_list.append(temp)
            elif executed_each_sprint and execution_phase:
                temp  = 'find_in_set("%s/%s", executed_each_sprint)' % (execution_phase, executed_each_sprint)
                where_list.append(temp)
            else:
                temp  = 'locate("/%s", executed_each_sprint)' % executed_each_sprint
                where_list.append(temp)

            if show_history_only == '1' and sprint_id:
                temp = 'sprint_id < %s OR sprint_id is null' % sprint_id
                where_list.append(temp)

            if creater:
                filters['creater_name'] = creater

            if updater:
                filters['updater_name'] = updater

            if tag:
                temp = 'find_in_set("%s", tags)' % tag
                where_list.append(temp)

            if recursive:
                suite_id_list = []
                suite_id_list.append(int(suite_id))

                # 获取子测试套件id
                query_sql = 'SELECT id FROM tb_testcase_suite AS t ' \
                           'WHERE FIND_IN_SET(%s, t.all_upper_node_ids) AND is_delete=0' % int(suite_id)
                query_result = TestcaseSuite.objects.raw(query_sql)
                sub_suite_id_list = []
                for item in query_result:
                   sub_suite_id_list.append(int(item.id))
                suite_id_list.extend(sub_suite_id_list)
            else:
                suite_id_list = [int(suite_id)]
            if suite_id_list:
                filters['suite_id__in'] = suite_id_list
            else:
                result['msg'] =  '导出用例失败，获取测试用所属测试套件为空'
                result['success'] =  False
                return Response(result, status.HTTP_200_OK)

            # 获取套件关联的测试用例
            if where_list:
                testcases = globals()[case_table_model].objects.filter(**filters).extra(where=where_list).order_by(*sort_list)[0:MAX_CASES_NUM_XMIND_EXPORT]
            else:
                testcases = globals()[case_table_model].objects.filter(**filters).order_by(*sort_list)[0:MAX_CASES_NUM_XMIND_EXPORT]

            temp_dict = OrderedDict()
            for testcase in testcases:
                suite_id = testcase.suite_id
                if suite_id not in temp_dict:
                    suite_path = TestSuiteStaticUtilsClass.get_suite_path(suite_id)
                    suite_path = '/' + suite_path
                    temp_dict[suite_id] = {
                        'suitePath':suite_path,
                        'children':[]
                    }
                testcase.suitePath = temp_dict[suite_id]['suitePath']
                row = {'suite_path': testcase.suitePath, 'name': testcase.name, 'priority':testcase.priority, 'desc':testcase.desc, 'steps':testcase.steps}
                temp_dict[suite_id]['children'].append(row)

            workbook = xmind.load(XMIND_TEMPLATE_FILE_PATH)
            xmind_node_dict = {} # 存放 xmind 结点和测试套件的映射关系
            for suite_id in temp_dict.keys():
                children = temp_dict[suite_id]['children']
                for row in children:
                    suite_path = row.get('suite_path')
                    if suite_path.count('/') < 2: # 不允许导出归属一级套件下的用例
                        continue
                    # XMindTestcaseExportAPIView.create_xmind_case_suite_topic(suite_path,xmind_node_dict,first_sheet)
                    XMindTestcaseExportAPIView.create_xmind_case_suite_topic(suite_path, xmind_node_dict, workbook)
                    steps = json.loads(row.get('steps'), object_pairs_hook=OrderedDict)
                    case_topic = xmind_node_dict[suite_path].addSubTopic()
                    case_topic.setTitle(row.get('name'))
                    desc = row.get('desc')
                    priority =  row.get('priority').lstrip('P')
                    case_topic.addMarker("priority-" + priority)
                    if desc:
                        case_topic.setPlainNotes(desc)

                    for step in steps:
                        result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', step['action'])
                        if result:
                            action = step['action'].lstrip(result[0]).strip()
                        else:
                            action = step['action']
                        result = re.findall('(^\d+?\s*[、|.|,|，| ]\s*)', step['expection'])
                        if result:
                            expection = step['expection'].lstrip(result[0]).strip()
                        else:
                            expection = step['expection']
                        action_topic = None
                        if action:
                            action_topic = case_topic.addSubTopic()
                            action_topic.setTitle(action)

                        if expection:
                            expection = expection.replace('\n\n', '\n')
                            if not action_topic:
                                action_topic = case_topic.addSubTopic()
                                action_topic.setTitle('步骤略')
                            expection_topic = action_topic.addSubTopic()
                            expection_topic.setTitle(expection)

            time_str = timezone.now().strftime('%Y%m%d%H%M%S%f')
            file_name =   shortuuid.uuid() + time_str + '.xmind'
            file_dir = settings.MEDIA_ROOT.rstrip('/') + '/testcase/download'
            if not os.path.exists(file_dir):# 路径不存在
                if not mkdirs_in_batch(file_dir):
                    result['msg'] =  '导出用例失败，批量创建路径(%s)对应的目录失败'  % file_dir
                    result['success'] =  False
                    return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
            file_absolute_path =  '%s/%s' % (file_dir, file_name)
            xmind.save(workbook=workbook, path=file_absolute_path)  # 不改动原始文件，另存为其它xmind文件，等同 xmind.save(workbook, 'd:\\example\\exam.xmind')

            file = open(file_absolute_path, 'rb')
            file_response = FileResponse(file)
            file_response['Content-Type']='application/octet-stream'
            temp_file_name = '导出测试用例%s' %  time_str
            file_response["Access-Control-Expose-Headers"] = 'Content-Disposition'
            file_response['Content-Disposition']='attachment;filename={}.xmind'.format(urlquote(temp_file_name))

            relative_file_path = '/testcase/download/%s' % file_name
            try:
                # 存储导出测试用例文件到数据库
                obj = NormalFileDownload(type='测试用例', name=temp_file_name, file_path=relative_file_path, creater=request.user.name, creater_id=request.user.id,  updater_id=request.user.id,  updater=request.user.name, is_delete=0)
                obj.save()
            except Exception as e:
                file.close()
                os.remove(file_absolute_path)
                raise Exception(e)
            return file_response
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)



class CopyTestcasesToSprintAPIView(APIView):
    '''
    测试视图-测试用例管理，克隆用例到迭代
    '''

    def post(self, request, format=None):
        result = {}
        try:
            data = request.data
            creater_id = request.user.id
            creater_name = request.user.name
            product_id = data.get('product_id')
            sprint_id = data.get('sprint_id')
            testcases = data.get('testcases') # 直接根据前端发来数据克隆，不走数据库

            # 获取复制到目标迭代项目对应的测试套件
            if not sprint_id:
                result['msg'] =  '克隆失败，克隆至目标迭代项目不能为空'
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)

            sprint_suite = TestcaseSuite.objects.filter(sprint_id = sprint_id, product_id=product_id, parent_id=-1, is_delete=0).first()
            if sprint_suite:
                # 遍历测试用例
                for testcase in testcases:
                    suite_parent_id =  sprint_suite.id
                    suite_type = testcase['suite_type']
                    if suite_type == 'base':
                       continue # 基线不让拷贝到迭代
                       copy_direction = 'base_to_sprint' # 拷贝方向,供复制用例步骤使用
                    elif suite_type == 'sprint':
                       copy_direction = 'sprint_to_sprint'
                    else:
                        result['msg'] =  '克隆失败，用例(guid:%s， 标题：%s)类型错误：%s' % (testcase['guid'], testcase['name'], suite_type)
                        result['success'] =  False
                        return Response(result, status.HTTP_400_BAD_REQUEST)

                    new_testcase = {}
                    new_testcase['guid'] = shortuuid.uuid()
                    new_testcase['custom_no'] = None

                    # 创建套件
                    suite_path = testcase['suite_path']
                    suite_name_list = suite_path.strip('/').split('/')
                    for suite_name in suite_name_list[1:]: # 第一层套件为根套件，要去掉，所以索引从1开始
                        # 查找父级测试套件下是否已存在相同名称的测试套件
                        sub_suite = TestcaseSuite.objects.filter(parent_id=suite_parent_id, name=suite_name, is_delete=0).first()
                        if sub_suite:
                            suite_parent_id = sub_suite.id
                        else:
                            all_upper_node_ids = TestSuiteStaticUtilsClass.get_father_suite_ids(suite_parent_id, [suite_parent_id])
                            all_upper_node_ids = ','.join([str(item) for item in all_upper_node_ids])
                            suite_data = {'name':suite_name, 'type': 'sprint', 'sprint_id':sprint_id, 'product_id': product_id, 'parent_id': suite_parent_id, 'creater_id': creater_id, 'updater_id':creater_id, 'is_delete':0,
                                          'all_upper_node_ids':all_upper_node_ids}
                            suite_data = TestSuiteStaticUtilsClass.create_testcase_suite(suite_data)
                            suite_parent_id = suite_data.get('id')

                    new_testcase['suite_id'] = suite_parent_id
                    new_testcase['product_id'] = testcase['product_id']
                    new_testcase['sprint_id'] = sprint_id
                    new_testcase['name'] = testcase['name']
                    new_testcase['priority'] = testcase['priority']
                    new_testcase['execution_phase'] = ','.join(testcase['execution_phase'])
                    new_testcase['execution_method'] = testcase['execution_method']
                    new_testcase['executed_each_sprint'] = testcase['executed_each_sprint']
                    new_testcase['tags'] = ','.join(testcase['tags'])
                    new_testcase['desc'] = testcase['desc']
                    new_testcase['precondition'] = testcase['precondition']
                    new_testcase['steps'] = testcase['steps']
                    new_testcase['postcondition'] = testcase['postcondition']
                    new_testcase['creater_id'] = creater_id
                    new_testcase['creater_name'] = creater_name
                    new_testcase['updater_id'] = creater_id
                    new_testcase['updater_name'] = creater_name
                    new_testcase['is_delete'] = 0

                    # 创建测试用例
                    try:
                        serializer = SprintTestcaseSerializer(data=new_testcase)
                        serializer.is_valid(raise_exception=True)
                        with transaction.atomic():
                            serializer.save()
                            to_case_guid = serializer.data.get('guid')
                            src_case_guid = testcase['guid']

                            # 拷贝源用例的测试附件到新建测试用例下
                            TestcaseAPIUtils.copy_case_attachments(to_case_guid, src_case_guid, copy_direction)

                            # 注释以下代码，因为基线用例不让复制到迭代了
                            # if suite_type == 'base': # 更新对应的基线用例
                            #     # 修改基线测试用例关联的迭代测试用例guid
                            #     base_case = BaseTestcase.objects.filter(guid=src_case_guid, is_delete=0).first()
                            #     if base_case:
                            #         base_case.sprint_testcase_guid = to_case_guid
                            #         base_case.save()
                            #     else:
                            #         # 小概率事件，被复制的基线用例被删除，啥也不做，
                            #         pass
                            # elif suite_type == 'sprint': # 同步创建基线测试用例
                            #     TestcaseAPIUtils.sync_sprint_case_to_base(to_case_guid, suite_parent_id, new_testcase)
                            base_case = BaseTestcase.objects.filter(sprint_testcase_guid=src_case_guid, is_delete=0).first()
                            if base_case:
                                base_case.sprint_testcase_guid = to_case_guid
                                base_case.save()
                            else:
                                # 被复制迭代用例对应的基线用例被删除，同步创建对应的基线用例
                                TestcaseAPIUtils.sync_sprint_case_to_base(to_case_guid, suite_parent_id, new_testcase)
                    except Exception as e:
                        result['msg'] =  '克隆失败：%s' % e
                        result['success'] =  False
                        return Response(result, status.HTTP_400_BAD_REQUEST)
            else:
                result['msg'] =  '克隆失败，找不到目标迭代项目关联的根测试集'
                result['success'] =  False
                return Response(result, status.HTTP_400_BAD_REQUEST)
            result['msg'] =  '克隆成功'
            result['success'] =  True
            return Response(result, status.HTTP_200_OK)
        except Exception as e:
            result['msg'] =  '%s' % e
            result['success'] =  False
            return Response(result, status.HTTP_500_INTERNAL_SERVER_ERROR)
